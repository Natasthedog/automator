from __future__ import annotations

import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook


class ProductDescriptionViewTests(TestCase):
    def _scope_upload(self, sheets: dict[str, list[str]]) -> SimpleUploadedFile:
        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)
        for sheet_name, headers in sheets.items():
            worksheet = workbook.create_sheet(title=sheet_name)
            worksheet.append(headers)
            worksheet.append(["value" for _ in headers])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            "scope.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _scope_upload_with_rows(self, sheets: dict[str, list[list[object]]]) -> SimpleUploadedFile:
        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)
        for sheet_name, rows in sheets.items():
            worksheet = workbook.create_sheet(title=sheet_name)
            for row in rows:
                worksheet.append(row)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            "scope.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_product_description_prompts_for_sheet_when_product_list_not_found(self):
        response = self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload(
                    {
                        "PRODUCT DESCRIPTION": ["Name", "Value"],
                        "LookupSheet": ["Manufacturer", "Brand", "Subbrand"],
                    }
                )
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Please choose the correct sheet from the list")
        self.assertContains(response, "LookupSheet")

    def test_product_description_renders_multi_column_rollup_builder_and_saves_rollup(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload(
                    {
                        "PRODUCT DESCRIPTION": ["Name", "Value"],
                        "Product List": ["Manufacturer", "Brand", "Subbrand"],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PRODUCT DESCRIPTION",
                "rollups": ["Manufacturer", "Brand"],
                "rollup_aliases": ["MFR", "BRAND"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Build roll ups")
        self.assertContains(response, "Add another roll up")
        self.assertContains(response, "repeat(3, minmax(0, 1fr))")
        self.assertContains(response, "choose column 2")
        self.assertContains(response, "choose column 3")
        self.assertContains(response, "Manufacturer")
        self.assertContains(response, "Brand")
        self.assertContains(response, "Rename roll up")
        self.assertContains(response, "Use Multi label")
        self.assertContains(response, "MFR")
        self.assertContains(response, "BRAND")
        self.assertContains(response, "Brand_Subbrand_Flavour")
        self.assertContains(response, "PPG_ID column")
        self.assertContains(response, "PPG_NAME column")
        self.assertContains(response, "PPG_EAN_CORRESPONDENCE sheet")
        self.assertContains(response, "Generate PRODUCT_DESCRIPTION & Download Scope")


    def test_save_rollups_warns_when_no_rollup_values_detected(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload(
                    {
                        "Product List": ["EAN", "Brand", "Subbrand"],
                        "PPG_EAN_CORRESPONDENCE": ["PPG_ID", "PPG_NAME", "EAN"],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "action": "save_rollups",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No roll ups were detected from your selections")



    def test_save_rollups_accepts_rollup_parts_payload(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload(
                    {
                        "Product List": ["EAN", "Brand", "Subbrand", "Flavour"],
                        "PPG_EAN_CORRESPONDENCE": ["PPG_ID", "PPG_NAME", "EAN"],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollup_part_1": ["Brand"],
                "rollup_part_2": ["Subbrand"],
                "rollup_part_3": ["Flavour"],
                "rollup_alias": ["ROL_1"],
                "action": "save_rollups",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Roll up selection saved for this session")
        self.assertContains(response, "Brand_Subbrand_Flavour")


    def test_generate_scope_downloads_product_description_sheet(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Manufacturer", "Brand", "PackType"],
                            ["111", "Mfr A", "Brand A", "Bottle"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN"],
                            ["P1", "PPG One", "111"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollups": ["Manufacturer", "PackType"],
                "rollup_aliases": ["MANUFACTURER", "PACK_TYPE"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            'attachment; filename="scope_with_product_description.xlsx"',
            response["Content-Disposition"],
        )



    def test_generate_scope_builds_three_column_rollup_values(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Brand", "Subbrand", "Flavour"],
                            ["111", "BrandA", "SubA", "Lemon"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN"],
                            ["P1", "PPG One", "111"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollups": ["Brand_Subbrand_Flavour"],
                "rollup_aliases": ["ROL1"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        loaded = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = loaded["PRODUCT_DESCRIPTION"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

        self.assertIn("ROL1", headers)
        self.assertIn("BrandA_SubA_Lemon", values)




    def test_generate_scope_builds_merged_rollup_when_source_columns_have_underscores(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Brand_Owner", "Sub_Brand", "Flavour"],
                            ["111", "OwnerA", "SubA", "Lemon"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN"],
                            ["P1", "PPG One", "111"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollup_part_1": ["Brand_Owner"],
                "rollup_part_2": ["Sub_Brand"],
                "rollup_part_3": ["Flavour"],
                "rollup_alias": ["ROL_US"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        loaded = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = loaded["PRODUCT_DESCRIPTION"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

        self.assertIn("ROL_US", headers)
        self.assertIn("OwnerA_SubA_Lemon", values)



    def test_generate_scope_joins_merged_rollup_values_with_underscores_not_spaces(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Brand Name", "Sub Brand", "Flavour Note"],
                            ["111", "Brand A", "Sub A", "Lemon Lime"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN"],
                            ["P1", "PPG One", "111"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollup_part_1": ["Brand Name"],
                "rollup_part_2": ["Sub Brand"],
                "rollup_part_3": ["Flavour Note"],
                "rollup_alias": ["ROL_SPACE"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        loaded = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = loaded["PRODUCT_DESCRIPTION"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

        self.assertIn("ROL_SPACE", headers)
        self.assertIn("Brand_A_Sub_A_Lemon_Lime", values)

    def test_generate_scope_joins_multiple_rollup_values_for_same_ppg(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Manufacturer", "Brand"],
                            ["111", "Manuf2", "BrandT"],
                            ["222", "Manuf2", "BrandZ"],
                            ["333", "Manuf2", "BrandX"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN"],
                            ["2", "NAME2", "111"],
                            ["2", "NAME2", "222"],
                            ["2", "NAME2", "333"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollups": ["Manufacturer", "Brand"],
                "rollup_aliases": ["MANUFACTURER", "BRAND"],
                "rollup_use_multi_label": ["0", "0"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        loaded = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = loaded["PRODUCT_DESCRIPTION"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        value_map = dict(zip(headers, values))

        self.assertEqual(value_map["PPG_ID"], "2")
        self.assertEqual(value_map["PPG_NAME"], "NAME2")
        self.assertEqual(value_map["MANUFACTURER"], "Manuf2")
        self.assertEqual(value_map["BRAND"], "BrandT_BrandZ_BrandX")


    def test_generate_scope_can_output_multi_label_for_selected_rollup(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Manufacturer", "Brand", "CVA"],
                            ["111", "MANUF2", "BrandT", "3000g"],
                            ["222", "MANUF4", "BrandZ", "5000g"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN"],
                            ["2", "NAME2", "111"],
                            ["2", "NAME2", "222"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollups": ["Manufacturer", "Brand", "CVA"],
                "rollup_aliases": ["MANUFACTURER", "BRAND", "CVA"],
                "rollup_use_multi_label": ["0", "1", "0"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        loaded = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = loaded["PRODUCT_DESCRIPTION"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        value_map = dict(zip(headers, values))

        self.assertEqual(value_map["MANUFACTURER"], "MANUF2_MANUF4")
        self.assertEqual(value_map["BRAND"], "Multi Brand")
        self.assertEqual(value_map["CVA"], "3000g_5000g")


    def test_generate_scope_handles_overlapping_column_names_between_sheets(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "CVA", "Brand"],
                            ["111", "Product CVA", "Brand A"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["PPG_ID", "PPG_NAME", "EAN", "CVA"],
                            ["P1", "PPG One", "111", "PPG CVA"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollups": ["CVA"],
                "rollup_aliases": ["CVA"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_generate_scope_warns_when_ppg_id_or_name_missing_and_not_selected(self):
        self.client.post(
            reverse("product-description"),
            data={
                "scope_workbook": self._scope_upload_with_rows(
                    {
                        "Product List": [
                            ["EAN", "Manufacturer"],
                            ["111", "Mfr A"],
                        ],
                        "PPG_EAN_CORRESPONDENCE": [
                            ["GROUP_ID", "GROUP_NAME", "EAN"],
                            ["P1", "PPG One", "111"],
                        ],
                    }
                )
            },
        )

        response = self.client.post(
            reverse("product-description"),
            data={
                "product_list_sheet": "Product List",
                "ppg_correspondence_sheet": "PPG_EAN_CORRESPONDENCE",
                "rollups": ["Manufacturer"],
                "rollup_aliases": ["MANUFACTURER"],
                "action": "generate_scope",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Please identify those columns before generating PRODUCT_DESCRIPTION")
