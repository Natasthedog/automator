import base64
import io
import logging
import re
from itertools import combinations

import pandas as pd
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import load_workbook

from .engine.waterfall.targets import _find_sheet_by_candidates

logger = logging.getLogger(__name__)
PROJECT_TEMPLATES = {}
PRODUCT_DESCRIPTION_SCOPE_KEY = "product_description_scope_workbook"
PRODUCT_DESCRIPTION_ROLLUPS_KEY = "product_description_rollups"
PRODUCT_DESCRIPTION_ROLLUP_ALIASES_KEY = "product_description_rollup_aliases"
PRODUCT_DESCRIPTION_ROLLUP_PARTS_KEY = "product_description_rollup_parts"
PRODUCT_DESCRIPTION_ROLLUP_MULTI_LABEL_KEY = "product_description_rollup_multi_label"


def _product_list_columns_from_sheet(workbook_bytes: bytes, sheet_name: str) -> list[str]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    for row in worksheet.iter_rows(values_only=True):
        values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if values:
            return values
    return []


def _normalized_text(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return "".join(ch for ch in text.lower() if ch.isalnum())


def _find_column_name(columns: list[str], target: str) -> str | None:
    normalized_target = _normalized_text(target)
    normalized_columns = {_normalized_text(column): column for column in columns}
    if normalized_target in normalized_columns:
        return normalized_columns[normalized_target]
    for normalized_column, column in normalized_columns.items():
        if normalized_target in normalized_column or normalized_column in normalized_target:
            return column
    return None


def _sheet_df_from_workbook(workbook_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(workbook_bytes), sheet_name=sheet_name, dtype=object)


def _underscore_joined_text(value: object) -> str:
    return re.sub(r"\s+", "_", str(value or "").strip())


def _joined_unique_values_by_group(group: pd.Series) -> list[str]:
    unique_values: list[str] = []
    seen: set[str] = set()
    for value in group:
        if pd.isna(value):
            continue
        normalized_value = _underscore_joined_text(value)
        if not normalized_value or normalized_value in seen:
            continue
        seen.add(normalized_value)
        unique_values.append(normalized_value)
    return unique_values


def _multi_rollup_label(label: str) -> str:
    cleaned = str(label or "").strip().replace("_", " ")
    return f"Multi {cleaned.title() if cleaned else 'Value'}"


def _is_multi_label_enabled(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() not in {"0", "false", "off", "no"}
    return bool(value)


def _build_product_description_df(
    correspondence_df: pd.DataFrame,
    product_list_df: pd.DataFrame,
    rollups: list[str],
    alias_map: dict[str, str],
    ppg_id_column: str,
    ppg_name_column: str,
    correspondence_mapping_column: str,
    product_list_mapping_column: str,
    rollup_parts_map: dict[str, list[str]] | None = None,
    rollup_multi_label_map: dict[str, bool] | None = None,
) -> tuple[pd.DataFrame, int]:
    right = product_list_df.copy()
    correspondence_columns = [ppg_id_column, ppg_name_column, correspondence_mapping_column]
    left = correspondence_df.loc[:, list(dict.fromkeys(correspondence_columns))].copy()

    left["__mapping_key"] = left[correspondence_mapping_column].map(_normalized_text)
    right["__mapping_key"] = right[product_list_mapping_column].map(_normalized_text)
    product_columns = [
        column for column in right.columns if column not in {"__mapping_key", product_list_mapping_column}
    ]

    df_ProductDescription = left.merge(
        right[["__mapping_key", *product_columns]],
        on="__mapping_key",
        how="left",
    )

    matched_rows = df_ProductDescription[df_ProductDescription["__mapping_key"] != ""]
    match_count = int((~matched_rows[product_columns].isna().all(axis=1)).sum()) if product_columns else 0

    selected_rollups: list[str] = []
    for rollup in rollups:
        component_columns = _rollup_component_columns(
            rollup,
            product_columns,
            (rollup_parts_map or {}).get(rollup),
        )
        if not component_columns:
            continue
        if len(component_columns) > 1:
            df_ProductDescription[rollup] = _compose_rollup_series(df_ProductDescription, component_columns)
        selected_rollups.append(rollup)

    matched_rows = df_ProductDescription[df_ProductDescription["__mapping_key"] != ""]
    match_count = int((~matched_rows[product_columns].isna().all(axis=1)).sum()) if product_columns else 0

    grouped = df_ProductDescription.groupby([ppg_id_column, ppg_name_column], dropna=False)[selected_rollups].agg(
        _joined_unique_values_by_group
    )

    for rollup in selected_rollups:
        use_multi_label = _is_multi_label_enabled((rollup_multi_label_map or {}).get(rollup), default=True)
        alias_label = alias_map.get(rollup, rollup)
        grouped[rollup] = grouped[rollup].map(
            lambda values: (
                _multi_rollup_label(alias_label)
                if use_multi_label and len(values) > 1
                else ("_".join(values) if values else None)
            )
        )

    grouped = grouped.reset_index()
    rename_map = {rollup: alias_map.get(rollup, rollup) for rollup in selected_rollups}
    grouped = grouped.rename(columns=rename_map)
    return grouped, match_count


def _rollup_component_columns(
    rollup: str,
    product_columns: list[str],
    submitted_parts: list[str] | None = None,
) -> list[str]:
    if submitted_parts:
        parts = [part.strip() for part in submitted_parts if part and part.strip()]
        if all(part in product_columns for part in parts):
            return parts
    if rollup in product_columns:
        return [rollup]
    parts = [part.strip() for part in (rollup or "").split("_") if part.strip()]
    if len(parts) < 2 or len(parts) > 3:
        return []
    if all(part in product_columns for part in parts):
        return parts
    return []


def _compose_rollup_series(df: pd.DataFrame, component_columns: list[str]) -> pd.Series:
    def _compose_row(row: pd.Series):
        values: list[str] = []
        for column in component_columns:
            value = row.get(column)
            if pd.notna(value) and str(value).strip():
                values.append(_underscore_joined_text(value))
        return "_".join(values) if values else None

    return df[component_columns].apply(_compose_row, axis=1)


def _updated_scope_workbook_bytes(
    workbook_bytes: bytes,
    sheet_name: str,
    product_description_df: pd.DataFrame,
) -> bytes:
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(list(product_description_df.columns))
    for row in product_description_df.itertuples(index=False):
        worksheet.append(list(row))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalized_rollups(rollups: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for rollup in rollups:
        value = (rollup or "").strip().strip("_")
        if not value or value in seen:
            continue
        seen.add(value)
        normalized.append(value)
    return normalized


def _rollup_alias_map(rollups: list[str], aliases: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for rollup, alias in zip(rollups, aliases):
        key = (rollup or "").strip()
        if not key:
            continue
        value = (alias or "").strip()
        result[key] = value or key
    return result


def _expanded_multi_label_flags(multi_label_flags: list[str]) -> list[str]:
    if len(multi_label_flags) != 1:
        return multi_label_flags
    raw = str(multi_label_flags[0] or "").strip()
    if not raw:
        return []
    tokens = re.findall(r"0|1|true|false|off|on|no|yes", raw.lower())
    if len(tokens) > 1:
        return tokens
    return multi_label_flags


def _rollup_multi_label_map(rollups: list[str], multi_label_flags: list[str]) -> dict[str, bool]:
    result: dict[str, bool] = {}
    expanded_flags = _expanded_multi_label_flags(multi_label_flags)
    for index, rollup in enumerate(rollups):
        key = (rollup or "").strip()
        if not key:
            continue
        flag = (expanded_flags[index] if index < len(expanded_flags) else "1").strip().lower()
        result[key] = _is_multi_label_enabled(flag, default=True)
    return result


def _rollups_from_submitted_parts(request) -> tuple[list[str], list[str], dict[str, list[str]], list[str]]:
    part_one = request.POST.getlist("rollup_part_1")
    part_two = request.POST.getlist("rollup_part_2")
    part_three = request.POST.getlist("rollup_part_3")
    aliases = request.POST.getlist("rollup_alias")
    multi_label_flags = request.POST.getlist("rollup_use_multi_label")

    max_rows = max(len(part_one), len(part_two), len(part_three), len(aliases), len(multi_label_flags), 0)
    rollups: list[str] = []
    rollup_aliases: list[str] = []
    rollup_parts_map: dict[str, list[str]] = {}
    rollup_multi_label_flags: list[str] = []
    for index in range(max_rows):
        parts = [
            (part_one[index] if index < len(part_one) else "").strip(),
            (part_two[index] if index < len(part_two) else "").strip(),
            (part_three[index] if index < len(part_three) else "").strip(),
        ]
        selected = [value for value in parts if value]
        if not selected:
            continue
        rollup = "_".join(_underscore_joined_text(value) for value in selected)
        alias = (aliases[index] if index < len(aliases) else "").strip() or rollup
        rollups.append(rollup)
        rollup_aliases.append(alias)
        rollup_parts_map[rollup] = selected
        rollup_multi_label_flags.append(multi_label_flags[index] if index < len(multi_label_flags) else "1")
    return _normalized_rollups(rollups), rollup_aliases, rollup_parts_map, rollup_multi_label_flags


def home(request):
    return redirect("file-uploads")


def product_description(request):
    context: dict[str, object] = {
        "sheet_names": [],
        "product_list_columns": [],
        "ppg_sheet_columns": [],
        "selected_rollups": request.session.get(PRODUCT_DESCRIPTION_ROLLUPS_KEY, []),
        "selected_rollup_aliases": request.session.get(PRODUCT_DESCRIPTION_ROLLUP_ALIASES_KEY, {}),
        "selected_sheet": "",
        "selected_ppg_sheet": "",
        "selected_product_list_mapping_column": "",
        "selected_ppg_mapping_column": "",
        "selected_ppg_id_column": "",
        "selected_ppg_name_column": "",
        "selected_rollup_pairs": [],
        "selected_rollup_parts": request.session.get(PRODUCT_DESCRIPTION_ROLLUP_PARTS_KEY, {}),
        "selected_rollup_multi_labels": request.session.get(PRODUCT_DESCRIPTION_ROLLUP_MULTI_LABEL_KEY, {}),
    }

    stored_scope = request.session.get(PRODUCT_DESCRIPTION_SCOPE_KEY)
    scope_bytes: bytes | None = None
    if stored_scope:
        scope_bytes = base64.b64decode(stored_scope)

    selected_rollups = context["selected_rollups"]
    selected_rollup_aliases = context["selected_rollup_aliases"]
    selected_rollup_multi_labels = context["selected_rollup_multi_labels"]
    context["selected_rollup_pairs"] = [
        {
            "value": rollup,
            "alias": selected_rollup_aliases.get(rollup, rollup),
            "use_multi_label": _is_multi_label_enabled(selected_rollup_multi_labels.get(rollup), default=True),
        }
        for rollup in selected_rollups
    ]

    if request.method == "POST":
        uploaded_scope = request.FILES.get("scope_workbook")
        selected_sheet = (request.POST.get("product_list_sheet") or "").strip()
        selected_ppg_sheet = (request.POST.get("ppg_correspondence_sheet") or "").strip()
        submitted_rollups_from_parts, submitted_aliases_from_parts, submitted_rollup_parts_map, submitted_multi_label_flags = _rollups_from_submitted_parts(request)
        submitted_rollups = submitted_rollups_from_parts or _normalized_rollups(request.POST.getlist("rollups"))
        submitted_aliases = submitted_aliases_from_parts or request.POST.getlist("rollup_aliases")
        submitted_multi_label_flags = submitted_multi_label_flags or request.POST.getlist("rollup_use_multi_label")
        selected_product_list_mapping_column = (
            request.POST.get("product_list_mapping_column") or ""
        ).strip()
        selected_ppg_mapping_column = (request.POST.get("ppg_mapping_column") or "").strip()
        selected_ppg_id_column = (request.POST.get("ppg_id_column") or "").strip()
        selected_ppg_name_column = (request.POST.get("ppg_name_column") or "").strip()
        action = (request.POST.get("action") or "").strip().lower()

        if uploaded_scope:
            scope_bytes = uploaded_scope.read()
            request.session[PRODUCT_DESCRIPTION_SCOPE_KEY] = base64.b64encode(scope_bytes).decode("ascii")
            request.session.modified = True

        if not scope_bytes:
            context["error"] = "Please upload a scope workbook to configure Product Description roll ups."
            return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

        workbook = load_workbook(io.BytesIO(scope_bytes), data_only=True, read_only=True)
        sheet_names = workbook.sheetnames
        context["sheet_names"] = sheet_names

        if not selected_sheet:
            selected_sheet = _find_sheet_by_candidates(sheet_names, "Product List") or ""
            if not selected_sheet:
                selected_sheet = _find_sheet_by_candidates(sheet_names, "ProductList") or ""
            if not selected_sheet:
                selected_sheet = _find_sheet_by_candidates(sheet_names, "Product_List") or ""

        if not selected_ppg_sheet:
            selected_ppg_sheet = _find_sheet_by_candidates(sheet_names, "PPG_EAN_CORRESPONDENCE") or ""

        context["selected_sheet"] = selected_sheet
        context["selected_ppg_sheet"] = selected_ppg_sheet

        if not selected_sheet:
            context["warning"] = (
                "We could not identify the Product List sheet automatically. "
                "Please choose the correct sheet from the list."
            )
            return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

        if selected_sheet not in sheet_names:
            context["error"] = "The selected Product List sheet does not exist in this workbook."
            return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

        if not selected_ppg_sheet:
            context["warning"] = (
                "We could not identify the PPG_EAN_CORRESPONDENCE sheet automatically. "
                "Please choose which sheet contains PPG_ID, PPG_NAME and EAN."
            )
            return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

        if selected_ppg_sheet not in sheet_names:
            context["error"] = "The selected PPG_EAN_CORRESPONDENCE sheet does not exist in this workbook."
            return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

        columns = _product_list_columns_from_sheet(scope_bytes, selected_sheet)
        ppg_columns = _product_list_columns_from_sheet(scope_bytes, selected_ppg_sheet)
        context["product_list_columns"] = columns
        context["ppg_sheet_columns"] = ppg_columns

        if submitted_rollups:
            alias_map = _rollup_alias_map(submitted_rollups, submitted_aliases)
            multi_label_map = _rollup_multi_label_map(submitted_rollups, submitted_multi_label_flags)
            request.session[PRODUCT_DESCRIPTION_ROLLUPS_KEY] = submitted_rollups
            request.session[PRODUCT_DESCRIPTION_ROLLUP_ALIASES_KEY] = alias_map
            request.session[PRODUCT_DESCRIPTION_ROLLUP_PARTS_KEY] = submitted_rollup_parts_map
            request.session[PRODUCT_DESCRIPTION_ROLLUP_MULTI_LABEL_KEY] = multi_label_map
            request.session.modified = True
            context["selected_rollups"] = submitted_rollups
            context["selected_rollup_aliases"] = alias_map
            context["selected_rollup_multi_labels"] = multi_label_map
            context["message"] = "Roll up selection saved for this session."
            context["selected_rollup_pairs"] = [
                {
                    "value": rollup,
                    "alias": alias_map.get(rollup, rollup),
                    "use_multi_label": _is_multi_label_enabled(multi_label_map.get(rollup), default=True),
                }
                for rollup in submitted_rollups
            ]
            context["selected_rollup_parts"] = submitted_rollup_parts_map
        elif action == "save_rollups":
            context["warning"] = (
                "No roll ups were detected from your selections. "
                "Choose 1 to 3 columns per row before saving."
            )

        context["selected_product_list_mapping_column"] = selected_product_list_mapping_column
        context["selected_ppg_mapping_column"] = selected_ppg_mapping_column
        context["selected_ppg_id_column"] = selected_ppg_id_column
        context["selected_ppg_name_column"] = selected_ppg_name_column

        if action == "generate_scope":
            saved_rollups = _normalized_rollups(request.session.get(PRODUCT_DESCRIPTION_ROLLUPS_KEY, []))
            if not saved_rollups:
                context["error"] = "Please save at least one roll up before generating the PRODUCT_DESCRIPTION sheet."
                return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

            alias_map = request.session.get(PRODUCT_DESCRIPTION_ROLLUP_ALIASES_KEY, {})
            saved_rollup_parts_map = request.session.get(PRODUCT_DESCRIPTION_ROLLUP_PARTS_KEY, {})
            saved_multi_label_map = request.session.get(PRODUCT_DESCRIPTION_ROLLUP_MULTI_LABEL_KEY, {})
            product_df = _sheet_df_from_workbook(scope_bytes, selected_sheet)
            correspondence_df = _sheet_df_from_workbook(scope_bytes, selected_ppg_sheet)

            ppg_id_column = selected_ppg_id_column or _find_column_name(list(correspondence_df.columns), "PPG_ID")
            ppg_name_column = selected_ppg_name_column or _find_column_name(list(correspondence_df.columns), "PPG_NAME")
            if not ppg_id_column or not ppg_name_column:
                context["warning"] = (
                    "Could not find PPG_ID and/or PPG_NAME columns in the selected PPG sheet. "
                    "Please identify those columns before generating PRODUCT_DESCRIPTION."
                )
                return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

            auto_ppg_mapping_column = _find_column_name(list(correspondence_df.columns), "EAN")
            correspondence_mapping_column = selected_ppg_mapping_column or auto_ppg_mapping_column
            if not correspondence_mapping_column:
                context["warning"] = (
                    "Could not find the EAN column in PPG_EAN_CORRESPONDENCE. "
                    "Please identify the mapping column in that sheet."
                )
                return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

            auto_product_mapping_column = _find_column_name(list(product_df.columns), "EAN")
            product_mapping_column = selected_product_list_mapping_column or auto_product_mapping_column
            if not product_mapping_column:
                context["warning"] = (
                    "Could not find the mapping column in Product List. "
                    "Please identify which Product List column maps to PPG_EAN_CORRESPONDENCE."
                )
                return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

            product_description_df, match_count = _build_product_description_df(
                correspondence_df,
                product_df,
                saved_rollups,
                alias_map,
                ppg_id_column,
                ppg_name_column,
                correspondence_mapping_column,
                product_mapping_column,
                saved_rollup_parts_map,
                saved_multi_label_map,
            )

            if match_count <= 0:
                context["warning"] = (
                    "Could not find matches between Product List and PPG_EAN_CORRESPONDENCE using selected mapping columns. "
                    "Please identify the correct mapping column(s)."
                )
                return render(request, "deck/PRODUCT_DESCRIPTION.html", context)

            updated_scope = _updated_scope_workbook_bytes(
                scope_bytes,
                "PRODUCT_DESCRIPTION",
                product_description_df,
            )
            response = HttpResponse(
                updated_scope,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="scope_with_product_description.xlsx"'
            return response

    return render(request, "deck/PRODUCT_DESCRIPTION.html", context)




def _find_column_by_candidates(df: pd.DataFrame, candidates: list[str]) -> str | None:
    column_lookup = {_normalized_text(column): column for column in df.columns}
    for candidate in candidates:
        key = _normalized_text(candidate)
        if key in column_lookup:
            return column_lookup[key]
    for candidate in candidates:
        key = _normalized_text(candidate)
        for normalized_column, original in column_lookup.items():
            if key and (key in normalized_column or normalized_column in key):
                return original
    return None


def _safe_group_correlation(group: pd.DataFrame, sales_col: str, bprv_col: str) -> float | None:
    clean = group[[sales_col, bprv_col]].apply(pd.to_numeric, errors="coerce").dropna()
    if len(clean) < 3:
        return None
    if clean[sales_col].std() < 1e-6 or clean[bprv_col].std() < 1e-6:
        return None
    return float(clean[sales_col].corr(clean[bprv_col]))


def _sales_spike_metrics(group: pd.DataFrame, sales_col: str, bprv_col: str) -> pd.Series:
    clean = group[[sales_col, bprv_col]].apply(pd.to_numeric, errors="coerce").dropna()
    if len(clean) < 4:
        return pd.Series({"slope": None, "spike_lift_pct": None, "spike_rows": 0})

    sales_variance = clean[sales_col].var()
    slope = None
    if sales_variance and sales_variance > 1e-12:
        slope = float(clean[sales_col].cov(clean[bprv_col]) / sales_variance)

    threshold = clean[sales_col].quantile(0.75)
    is_spike = clean[sales_col] >= threshold
    spike_rows = int(is_spike.sum())
    non_spike_rows = int((~is_spike).sum())
    if spike_rows < 2 or non_spike_rows < 2:
        return pd.Series({"slope": slope, "spike_lift_pct": None, "spike_rows": spike_rows})

    spike_mean = clean.loc[is_spike, bprv_col].mean()
    non_spike_mean = clean.loc[~is_spike, bprv_col].mean()
    if pd.isna(non_spike_mean) or abs(non_spike_mean) < 1e-12:
        spike_lift_pct = None
    else:
        spike_lift_pct = float(((spike_mean - non_spike_mean) / abs(non_spike_mean)) * 100)

    return pd.Series({"slope": slope, "spike_lift_pct": spike_lift_pct, "spike_rows": spike_rows})


def preqc_bprv(request):
    context: dict[str, object] = {
        "top_results": [],
        "chart_entries": [],
        "correlation_intro": "",
        "scope_sheet_options": [],
        "scope_columns": [],
        "bprv_columns": [],
        "selected_scope_sheet": "",
        "selected_scope_ppg_column": "",
        "selected_scope_brand_column": "",
        "selected_bprv_ppg_column": "",
        "selected_bprv_brand_column": "",
        "selected_bprv_geography_column": "",
        "selected_bprv_sales_column": "",
        "selected_bprv_value_column": "",
        "selected_bprv_week_column": "",
    }

    scope_key = "preqc_bprv_scope_workbook"
    bprv_key = "preqc_bprv_bprv_workbook"

    if request.method == "POST":
        scope_upload = request.FILES.get("scope_workbook")
        bprv_upload = request.FILES.get("bprv_workbook")
        if scope_upload:
            request.session[scope_key] = base64.b64encode(scope_upload.read()).decode("ascii")
            request.session.modified = True
        if bprv_upload:
            request.session[bprv_key] = base64.b64encode(bprv_upload.read()).decode("ascii")
            request.session.modified = True

        scope_encoded = request.session.get(scope_key)
        bprv_encoded = request.session.get(bprv_key)
        if not scope_encoded or not bprv_encoded:
            context["error"] = "Please upload both the scope file and the BPRV file."
            return render(request, "deck/PREQC_BPRV.html", context)

        scope_bytes = base64.b64decode(scope_encoded)
        bprv_bytes = base64.b64decode(bprv_encoded)

        try:
            scope_workbook = load_workbook(io.BytesIO(scope_bytes), data_only=True, read_only=True)
            scope_sheet_names = list(scope_workbook.sheetnames)
            context["scope_sheet_options"] = scope_sheet_names
        except Exception:
            context["error"] = "Could not read the scope workbook."
            return render(request, "deck/PREQC_BPRV.html", context)

        try:
            bprv_df = pd.read_excel(io.BytesIO(bprv_bytes), dtype=object)
            context["bprv_columns"] = list(bprv_df.columns)
        except Exception:
            context["error"] = "Could not read the BPRV workbook. Please verify the file format."
            return render(request, "deck/PREQC_BPRV.html", context)

        selected_scope_sheet = (request.POST.get("scope_product_description_sheet") or "").strip()
        auto_scope_sheet = _find_sheet_by_candidates(scope_sheet_names, ["PRODUCT_DESCRIPTION", "Product Description"])
        scope_sheet = selected_scope_sheet or auto_scope_sheet
        context["selected_scope_sheet"] = scope_sheet or ""

        if not scope_sheet:
            context["warning"] = "Could not find the PRODUCT DESCRIPTION sheet automatically. Please identify it."
            return render(request, "deck/PREQC_BPRV.html", context)

        try:
            scope_product_df = pd.read_excel(io.BytesIO(scope_bytes), sheet_name=scope_sheet, dtype=object)
            context["scope_columns"] = list(scope_product_df.columns)
        except Exception:
            context["error"] = "Could not read the selected sheet from the scope workbook."
            return render(request, "deck/PREQC_BPRV.html", context)

        selected_bprv_ppg_column = (request.POST.get("bprv_ppg_column") or "").strip()
        selected_bprv_brand_column = (request.POST.get("bprv_brand_column") or "").strip()
        selected_bprv_geography_column = (request.POST.get("bprv_geography_column") or "").strip()
        selected_bprv_sales_column = (request.POST.get("bprv_sales_column") or "").strip()
        selected_bprv_value_column = (request.POST.get("bprv_value_column") or "").strip()
        selected_bprv_week_column = (request.POST.get("bprv_week_column") or "").strip()
        selected_scope_ppg_column = (request.POST.get("scope_ppg_column") or "").strip()
        selected_scope_brand_column = (request.POST.get("scope_brand_column") or "").strip()

        context["selected_bprv_ppg_column"] = selected_bprv_ppg_column
        context["selected_bprv_brand_column"] = selected_bprv_brand_column
        context["selected_bprv_geography_column"] = selected_bprv_geography_column
        context["selected_bprv_sales_column"] = selected_bprv_sales_column
        context["selected_bprv_value_column"] = selected_bprv_value_column
        context["selected_bprv_week_column"] = selected_bprv_week_column
        context["selected_scope_ppg_column"] = selected_scope_ppg_column
        context["selected_scope_brand_column"] = selected_scope_brand_column

        bprv_ppg_col = selected_bprv_ppg_column or _find_column_by_candidates(bprv_df, ["PPG", "PPG_NAME", "PPG Name"])
        bprv_brand_col = selected_bprv_brand_column or _find_column_by_candidates(bprv_df, ["Brand"])
        geography_col = selected_bprv_geography_column or _find_column_by_candidates(
            bprv_df, ["Geography", "Area", "Region", "Market"]
        )
        sales_col = selected_bprv_sales_column or _find_column_by_candidates(bprv_df, ["Sales", "Value Sales", "Net Sales"])
        bprv_value_col = selected_bprv_value_column or _find_column_by_candidates(bprv_df, ["BPRV", "Bprv"])
        week_col = selected_bprv_week_column or _find_column_by_candidates(
            bprv_df, ["Week", "Weeks", "Date", "Period", "Company Week", "Company_Week", "Days"]
        )

        if not bprv_ppg_col:
            context["warning"] = "Could not find the PPG column in the BPRV file. Please identify it."
            return render(request, "deck/PREQC_BPRV.html", context)

        if not geography_col or not sales_col or not bprv_value_col:
            context["warning"] = (
                "Could not detect Geography, Sales, or BPRV column in the BPRV file. Please identify those columns."
            )
            return render(request, "deck/PREQC_BPRV.html", context)

        rename_map = {
            bprv_ppg_col: "PPG",
            geography_col: "Geography",
            sales_col: "Sales",
            bprv_value_col: "BPRV",
        }
        if week_col:
            rename_map[week_col] = "Week"

        merged = bprv_df.rename(columns=rename_map).copy()

        if bprv_brand_col:
            merged = merged.rename(columns={bprv_brand_col: "Brand"})
        else:
            scope_ppg_col = selected_scope_ppg_column or _find_column_by_candidates(
                scope_product_df, ["PPG", "PPG_NAME", "PPG Name"]
            )
            if not scope_ppg_col:
                context["warning"] = (
                    "Could not find the PPG column in the selected scope sheet. Please identify it."
                )
                return render(request, "deck/PREQC_BPRV.html", context)

            scope_brand_col = selected_scope_brand_column or _find_column_by_candidates(scope_product_df, ["Brand"])
            if not scope_brand_col:
                context["warning"] = (
                    "Could not find the Brand column in the selected scope sheet. Please identify it."
                )
                return render(request, "deck/PREQC_BPRV.html", context)

            context["selected_scope_ppg_column"] = scope_ppg_col
            context["selected_scope_brand_column"] = scope_brand_col
            brand_map = (
                scope_product_df[[scope_ppg_col, scope_brand_col]]
                .rename(columns={scope_ppg_col: "PPG", scope_brand_col: "Brand"})
                .dropna(subset=["PPG", "Brand"])
            )
            brand_map["__ppg_key"] = brand_map["PPG"].map(_normalized_text)
            brand_map = brand_map.drop_duplicates("__ppg_key")

            merged["__ppg_key"] = merged["PPG"].map(_normalized_text)
            merged = merged.merge(brand_map[["__ppg_key", "Brand"]], on="__ppg_key", how="left")

        merged = merged.dropna(subset=["PPG", "Brand"])

        grouped = (
            merged.groupby(["Geography", "PPG"], dropna=False)
            .apply(
                lambda g: pd.concat(
                    [
                        pd.Series(
                            {
                                "correlation": _safe_group_correlation(g, "Sales", "BPRV"),
                                "total_sales": pd.to_numeric(g["Sales"], errors="coerce").sum(),
                                "n_rows": len(g),
                            }
                        ),
                        _sales_spike_metrics(g, "Sales", "BPRV"),
                    ]
                )
            )
            .reset_index()
            .dropna(subset=["correlation"])
            .sort_values(by=["correlation", "total_sales"], ascending=[False, False])
        )

        grouped = grouped.sort_values(
            by=["total_sales", "correlation", "spike_lift_pct", "n_rows", "slope"],
            ascending=[False, False, False, False, False],
        )

        top_results = grouped.head(20).copy()
        top_results["correlation"] = pd.to_numeric(top_results["correlation"], errors="coerce").round(3)
        top_results["slope"] = pd.to_numeric(top_results["slope"], errors="coerce").round(4)
        top_results["spike_lift_pct"] = pd.to_numeric(top_results["spike_lift_pct"], errors="coerce").round(2)

        chart_entries: list[dict[str, object]] = []
        for _, row in top_results.iterrows():
            subset = merged[(merged["Geography"] == row["Geography"]) & (merged["PPG"] == row["PPG"])].copy()
            if subset.empty:
                continue
            if "Week" not in subset.columns:
                subset["Week"] = [f"W{i+1}" for i in range(len(subset))]
            else:
                as_dt = pd.to_datetime(subset["Week"], errors="coerce")
                if as_dt.notna().sum() >= max(2, len(subset) // 2):
                    subset["__week_sort"] = as_dt
                    subset = subset.sort_values("__week_sort")
                    subset["Week"] = subset["__week_sort"].dt.strftime("%Y-%m-%d")
                else:
                    subset["Week"] = subset["Week"].astype(str)

            subset["Sales"] = pd.to_numeric(subset["Sales"], errors="coerce")
            subset["BPRV"] = pd.to_numeric(subset["BPRV"], errors="coerce")
            subset = subset.dropna(subset=["Sales", "BPRV"]).reset_index(drop=True)
            if subset.empty:
                continue

            width, height = 420, 180
            pad_left, pad_right, pad_top, pad_bottom = 40, 40, 15, 30
            plot_w = width - pad_left - pad_right
            plot_h = height - pad_top - pad_bottom

            def _points(values: list[float]) -> str:
                if len(values) == 1:
                    xs = [pad_left + plot_w / 2]
                else:
                    xs = [pad_left + (plot_w * i / (len(values) - 1)) for i in range(len(values))]
                v_min = min(values)
                v_max = max(values)
                if abs(v_max - v_min) < 1e-12:
                    ys = [pad_top + plot_h / 2 for _ in values]
                else:
                    ys = [pad_top + plot_h - ((v - v_min) / (v_max - v_min) * plot_h) for v in values]
                return " ".join(f"{x:.2f},{y:.2f}" for x, y in zip(xs, ys))

            sales_values = subset["Sales"].tolist()
            bprv_values = subset["BPRV"].tolist()
            chart_entries.append(
                {
                    "geography": row["Geography"],
                    "ppg": row["PPG"],
                    "week_start": subset["Week"].iloc[0],
                    "week_end": subset["Week"].iloc[-1],
                    "sales_line": _points(sales_values),
                    "bprv_line": _points(bprv_values),
                    "sales_min": round(float(min(sales_values)), 2),
                    "sales_max": round(float(max(sales_values)), 2),
                    "bprv_min": round(float(min(bprv_values)), 2),
                    "bprv_max": round(float(max(bprv_values)), 2),
                }
            )

        report_buffer = io.BytesIO()
        with pd.ExcelWriter(report_buffer, engine="openpyxl") as writer:
            grouped.to_excel(writer, index=False, sheet_name="PreQC_BPRV_Correlation")
        report_bytes = report_buffer.getvalue()

        context["top_results"] = top_results.to_dict("records")
        context["chart_entries"] = chart_entries
        context["correlation_intro"] = (
            "Here we check the correlation between Sales and BPRV and rank the top PPG + Geography "
            "combinations by Total Sales, Correlation, Spike Lift, Active Weeks, and Slope."
        )
        request.session["preqc_bprv_report"] = base64.b64encode(report_bytes).decode("ascii")
        request.session.modified = True

    elif request.method == "GET" and (request.GET.get("action") or "").lower() == "download_report":
        encoded = request.session.get("preqc_bprv_report")
        if encoded:
            response = HttpResponse(
                base64.b64decode(encoded),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="preqc_bprv_full_report.xlsx"'
            return response

    return render(request, "deck/PREQC_BPRV.html", context)
def generate_deck(
    n_clicks,
    data_contents,
    data_name,
    scope_contents,
    scope_name,
    project_name,
    waterfall_targets,
    bucket_data,
):
    from dash import dcc, no_update

    from .engine.build import build_pptx_from_template
    from .engine.io_readers import (
        df_from_contents,
        product_description_df_from_contents,
        project_details_df_from_contents,
        scope_df_from_contents,
    )
    from .engine.project_details import _project_detail_value_from_df
    from .engine.waterfall.targets import target_brand_from_scope_df

    if not data_contents or not project_name:
        return no_update, "Please upload the data file and select a project."

    template_path = PROJECT_TEMPLATES.get(project_name)
    if not template_path or not template_path.exists():
        return no_update, "The selected project template could not be found."
    try:
        df = df_from_contents(data_contents, data_name)
        scope_df = None
        product_description_df = None
        project_details_df = None
        modelled_in_value = None
        metric_value = None
        if scope_contents:
            try:
                scope_df = scope_df_from_contents(scope_contents, scope_name)
            except Exception:
                scope_df = None
            try:
                product_description_df = product_description_df_from_contents(
                    scope_contents, scope_name
                )
            except Exception:
                product_description_df = None
            try:
                project_details_df = project_details_df_from_contents(
                    scope_contents, scope_name
                )
            except Exception:
                project_details_df = None
        if project_details_df is not None:
            modelled_in_value = _project_detail_value_from_df(
                project_details_df,
                "modelled in",
                [
                    "Sales will be modelled in",
                    "Sales will be modeled in",
                    "Sales modelled in",
                    "Sales modeled in",
                ],
                "Sales will be modelled in",
            )
            metric_value = _project_detail_value_from_df(
                project_details_df,
                "metric",
                [
                    "Volume metric (unique per dataset)",
                    "Volume metric unique per dataset",
                    "Volume metric",
                ],
                "Volume metric (unique per dataset)",
            )
        target_brand = target_brand_from_scope_df(scope_df)
        template_bytes = template_path.read_bytes()

        pptx_bytes = build_pptx_from_template(
            template_bytes,
            df,
            target_brand,
            project_name,
            scope_df,
            product_description_df,
            waterfall_targets,
            bucket_data,
            modelled_in_value,
            metric_value,
        )
        return dcc.send_bytes(lambda buff: buff.write(pptx_bytes), "deck.pptx"), "Building deck..."

    except Exception as exc:
        logger.exception("Deck generation failed.")
        message = str(exc).strip()
        if not message:
            message = "Unknown error. Check server logs for details."
        return no_update, f"Error ({type(exc).__name__}): {message}"


def download_waterfall_payloads(
    n_clicks,
    data_contents,
    data_name,
    scope_contents,
    scope_name,
    project_name,
    waterfall_targets,
    bucket_data,
):
    import io

    from dash import dcc, no_update
    from pptx import Presentation

    from .engine.io_readers import df_from_contents, scope_df_from_contents
    from .engine.pptx.charts import _waterfall_chart_from_slide
    from .engine.waterfall.compute import compute_waterfall_payloads_for_all_labels
    from .engine.waterfall.inject import _available_waterfall_template_slides
    from .engine.waterfall.payloads import _waterfall_payloads_to_json

    if not data_contents or not project_name:
        return no_update, "Please upload the data file and select a project."

    template_path = PROJECT_TEMPLATES.get(project_name)
    if not template_path or not template_path.exists():
        return no_update, "The selected project template could not be found."

    try:
        gathered_df = df_from_contents(data_contents, data_name)
        scope_df = None
        if scope_contents:
            try:
                scope_df = scope_df_from_contents(scope_contents, scope_name)
            except Exception:
                scope_df = None

        prs = Presentation(io.BytesIO(template_path.read_bytes()))
        available_slides = _available_waterfall_template_slides(prs)
        if not available_slides:
            return no_update, "Could not find the <Waterfall Template> slide in the template."
        template_chart = _waterfall_chart_from_slide(
            available_slides[0][1],
            "Waterfall Template",
        )
        if template_chart is None:
            return no_update, "Could not find the waterfall chart on the <Waterfall Template> slide."

        payloads_by_label = compute_waterfall_payloads_for_all_labels(
            gathered_df,
            scope_df,
            bucket_data,
            template_chart,
            target_labels=waterfall_targets,
        )
        payload_json = _waterfall_payloads_to_json(payloads_by_label)
        return dcc.send_string(payload_json, "waterfall_payloads.json"), "Prepared waterfall payload JSON."
    except Exception as exc:
        logger.exception("Waterfall payload JSON generation failed.")
        message = str(exc).strip()
        if not message:
            message = "Unknown error. Check server logs for details."
        return no_update, f"Error ({type(exc).__name__}): {message}"
